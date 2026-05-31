"""Parse Rosstat weekly CPI price indices and CPI basket weights from local XLSX files."""

import re
from datetime import date
from pathlib import Path

import openpyxl
import polars as pl

_RU_MONTHS: dict[str, int] = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}

# Strings that mark footnote / legend rows in prices sheets
_PRICE_FOOTNOTE_PREFIXES = ("*", "…", "1)", "2)")


def _parse_price_date(raw: str, year: int) -> date | None:
    """Parse a Rosstat date header like 'на 10 января **' or '26 декабря' into a date."""
    clean = re.sub(r"\*+", "", raw).strip()
    if clean.startswith("на "):
        clean = clean[3:].strip()
    parts = clean.split()
    if len(parts) != 2:
        return None
    month = _RU_MONTHS.get(parts[1])
    if month is None:
        return None
    try:
        return date(year, month, int(parts[0]))
    except (ValueError, TypeError):
        return None


def _is_price_footnote(value: str) -> bool:
    return any(value.startswith(p) for p in _PRICE_FOOTNOTE_PREFIXES)


def fetch_prices(prices_file: Path) -> pl.DataFrame:
    """Parse all year sheets from the Rosstat weekly CPI XLSX file.

    Returns a long-format DataFrame:
      date        – observation date
      category    – product / service name
      price_index – weekly price index, % of previous registration date
    """
    wb = openpyxl.load_workbook(prices_file, data_only=True)
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]

    records: list[dict] = []
    for sheet_name in year_sheets:
        year = int(sheet_name)
        ws = wb[sheet_name]

        # Row 4: col A = "Наименование", cols B+ = date header strings
        header = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        dates: list[date | None] = [
            _parse_price_date(str(v), year) if v is not None else None
            for v in header[1:]  # skip col A
        ]

        for row in ws.iter_rows(min_row=5, values_only=True):
            category = row[0]
            if category is None:
                continue
            category_str = str(category).strip()
            if _is_price_footnote(category_str):
                continue
            for i, value in enumerate(row[1:]):
                if i >= len(dates) or dates[i] is None or value is None:
                    continue
                try:
                    price_index = float(value)
                except (ValueError, TypeError):
                    continue  # skip "…" / text placeholders for missing data
                records.append({
                    "date": dates[i],
                    "category": category_str,
                    "price_index": price_index,
                })

    return (
        pl.DataFrame(
            records,
            schema={"date": pl.Date, "category": pl.String, "price_index": pl.Float64},
        )
        .sort(["date", "category"])
    )


def _detect_weight_columns(ws) -> tuple[int, int]:
    """Infer the (name_col, weight_col) 1-indexed positions from the sheet data.

    The weights file mixes three column layouts across years — 2004-2016 merge
    code+name into one column (weight in col 3), 2017-2023 prepend an "N п/п"
    column (weight in col 4), and 2024+ drop it again (weight in col 3) — and a
    couple of header rows (2013, 2023) are blank. Hard-coding positions silently
    drops the shifted years, so columns are inferred from the data instead:

      weight_col — the column whose values are fractional shares (0 < x < 100,
                   non-integer); product weights are fractional, codes and the
                   "N п/п" counter are integers.
      name_col   — the column carrying the longest text (product names), as
                   opposed to short numeric codes.
    """
    max_col = ws.max_column or 1
    frac_hits = {c: 0 for c in range(1, max_col + 1)}
    text_len = {c: 0 for c in range(1, max_col + 1)}
    for row in ws.iter_rows(min_row=11, max_row=40, values_only=True):
        for i, value in enumerate(row[:max_col], start=1):
            if isinstance(value, bool):
                continue
            if isinstance(value, (int, float)) and 0 < value < 100 and float(value) != int(value):
                frac_hits[i] += 1
            elif isinstance(value, str):
                clean = value.strip()
                if len(clean) > 3 and not clean.replace(".", "").replace(",", "").isdigit():
                    text_len[i] += len(clean)
    weight_col = max(frac_hits, key=frac_hits.get)
    name_col = max(text_len, key=text_len.get)
    return name_col, weight_col


def fetch_weights(weights_file: Path) -> pl.DataFrame:
    """Parse all year sheets from the Rosstat CPI basket weights XLSX file.

    Column positions vary by year, so they are inferred per sheet via
    _detect_weight_columns rather than hard-coded.

    Returns:
      year     – basket year
      code     – Rosstat product code (nullable; best-effort, unused downstream)
      category – product / service name
      weight   – share in CPI basket, % of total
    """
    wb = openpyxl.load_workbook(weights_file, data_only=True)
    year_sheets = [s for s in wb.sheetnames if s.isdigit()]

    records: list[dict] = []
    for sheet_name in year_sheets:
        year = int(sheet_name)
        ws = wb[sheet_name]
        name_col, weight_col = _detect_weight_columns(ws)
        code_col = name_col - 1 if name_col > 1 else None

        # Data starts at row 10 (row 10 is the "Все товары и услуги" total).
        for row in ws.iter_rows(min_row=10, values_only=True):
            if len(row) < weight_col:
                continue
            weight = row[weight_col - 1]
            name = row[name_col - 1]
            if weight is None or isinstance(weight, bool) or not isinstance(weight, (int, float)):
                continue
            if name is None:
                continue
            name_str = str(name).strip()
            if not name_str:
                continue
            code = row[code_col - 1] if code_col else None
            records.append({
                "year": year,
                "code": str(code) if code is not None else None,
                "category": name_str,
                "weight": float(weight),
            })

    return (
        pl.DataFrame(
            records,
            schema={
                "year": pl.Int32,
                "code": pl.String,
                "category": pl.String,
                "weight": pl.Float64,
            },
        )
        .sort(["year", "code"])
    )


def save_parquet(df: pl.DataFrame, output_dir: Path, filename: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / filename
    df.write_parquet(path)
    return path
